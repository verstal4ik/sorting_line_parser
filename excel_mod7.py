from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import re
import os

""" Словари объемов, последние цифры на конце. В словарь key - диаметр, value - объем в кубических метрах
27 для длин 2.7
3 для длин 3
57 для длин 57
6 для длин 6 
"""
volume27 = {10: 0.024, 11: 0.029, 12: 0.034, 13: 0.04, 14: 0.047, 15: 0.054, 16: 0.061, 17: 0.069, 18: 0.077, 19: 0.086, 20: 0.095, 21: 0.105, 22: 0.116, 23: 0.127, 24: 0.14, 25: 0.152, 26: 0.166, 27: 0.179, 28: 0.194, 29: 0.208, 30: 0.22, 31: 0.235, 32: 0.25, 33: 0.266, 34: 0.28, 35: 0.297, 36: 0.32, 37: 0.338, 38: 0.35, 39: 0.369, 40: 0.36, 41: 0.378, 42: 0.42, 43: 0.44, 44: 0.46, 45: 0.481, 46: 0.51, 47: 0.532, 48: 0.55, 49: 0.573, 50: 0.6, 51: 0.63, 52: 0.66, 53: 0.68, 54: 0.72, 55: 0.75, 56: 0.78, 57: 0.8, 58: 0.83, 59: 0.86, 60: 0.89, 61: 0.92, 62: 0.95, 63: 0.99, 64: 1.02, 65: 1.05, 66: 1.08, 67: 1.11, 68: 1.14, 69: 1.18, 70: 1.21}
volume3 = {10: 0.026, 11: 0.032, 12: 0.038, 13: 0.045, 14: 0.052, 15: 0.06, 16: 0.069, 17: 0.078, 18: 0.086, 19: 0.096, 20: 0.107, 21: 0.118, 22: 0.13, 23: 0.142, 24: 0.157, 25: 0.17, 26: 0.185, 27: 0.2, 28: 0.22, 29: 0.236, 30: 0.25, 31: 0.267, 32: 0.28, 33: 0.298, 34: 0.32, 35: 0.339, 36: 0.36, 37: 0.38, 38: 0.39, 39: 0.411, 40: 0.43, 41: 0.452, 42: 0.47, 43: 0.493, 44: 0.52, 45: 0.544, 46: 0.57, 47: 0.595, 48: 0.62, 49: 0.646, 50: 0.67, 51: 0.7, 52: 0.73, 53: 0.765, 54: 0.8, 55: 0.83, 56: 0.86, 57: 0.89, 58: 0.92, 59: 0.955, 60: 0.99, 61: 1.025, 62: 1.06, 63: 1.095, 64: 1.13, 65: 1.165, 66: 1.2, 67: 1.235, 68: 1.27, 69: 1.305, 70: 1.34}
volume57 = {10: 0.061, 11: 0.074, 12: 0.087, 13: 0.101, 14: 0.115, 15: 0.13, 16: 0.146, 17: 0.165, 18: 0.183, 19: 0.2, 20: 0.22, 21: 0.24, 22: 0.26, 23: 0.29, 24: 0.31, 25: 0.34, 26: 0.37, 27: 0.4, 28: 0.42, 29: 0.46, 30: 0.49, 31: 0.52, 32: 0.55, 33: 0.59, 34: 0.62, 35: 0.66, 36: 0.7, 37: 0.73, 38: 0.77, 39: 0.81, 40: 0.85, 41: 0.9, 42: 0.94, 43: 0.99, 44: 1.03, 45: 1.08, 46: 1.13, 47: 1.18, 48: 1.22, 49: 1.28, 50: 1.33, 51: 1.38, 52: 1.44, 53: 1.5, 54: 1.56, 55: 1.62, 56: 1.68, 57: 1.74, 58: 1.81, 59: 1.87, 60: 1.94, 61: 2.01, 62: 2.07, 63: 2.13, 64: 2.19, 65: 2.25, 66: 2.32, 67: 2.38, 68: 2.45, 69: 2.51, 70: 2.58} 
volume6 = {10: 0.065, 11: 0.08, 12: 0.093, 13: 0.108, 14: 0.123, 15: 0.139, 16: 0.155, 17: 0.174, 18: 0.194, 19: 0.212, 20: 0.23, 21: 0.255, 22: 0.28, 23: 0.305, 24: 0.33, 25: 0.36, 26: 0.39, 27: 0.42, 28: 0.45, 29: 0.485, 30: 0.52, 31: 0.555, 32: 0.59, 33: 0.625, 34: 0.66, 35: 0.7, 36: 0.74, 37: 0.78, 38: 0.82, 39: 0.86, 40: 0.9, 41: 0.95, 42: 1, 43: 1.045, 44: 1.09, 45: 1.14, 46: 1.19, 47: 1.245, 48: 1.3, 49: 1.355, 50: 1.41, 51: 1.47, 52: 1.53, 53: 1.59, 54: 1.65, 55: 1.715, 56: 1.78, 57: 1.845, 58: 1.91, 59: 1.98, 60: 2.05, 61: 2.115, 62: 2.18, 63: 2.25, 64: 2.32, 65: 2.38, 66: 2.44, 67: 2.505, 68: 2.57, 69: 2.645, 70: 2.72}

def main(timbers, group = [18, 20, 22, 24, 26, 35], mode=6, L1=5.7, L2=6.05, L3=6.20):
    """Функция сортировки списка бревен. 
    На входе массив из бревен timbers
    groups группы по которым надо разбить
    mode если указано 6, то складываем из 3 метровых 6 метровые . Если 3, то только
    p - для 6 метров = 2 , для 3 метров 1
    
    timbers6m - список бревен в зачет как 6 метров ["Диапазон", кол-во, объем] с итогом ['Итого', Сумма всех кол-в, Сумма всеx объемов]
    timbers57m - список бревен в зачет как 5.7 метров разбитый по группам как и timbers6m, с двумя итогам: до 6.05 и выше 6.20.
    timbers_def - список учтенный, как брак. Все браки из списка header_def сопоставленный с качеством указанных в sort_mathcing. Ключи сформированны в lsort, value: указатель в под каким индексом в regsort
    timbers1C - список для импорта в 1С
    timbers_logs - для фиксации log'a 
    timbers_data - список для вкладки данные, без изменений то, что было на входе.

    Diametr_KM - диаметр комля на окорке для автокачества
 
    L1 - первая граница длины
    L2 - вторая граница длины
    L3 - 3ья граница длины
       
    L1 < L2 < Norm L > L3 > L4 

    regsort - кортеж зарегистрированных сортов, которые формирует lsort
    regsort важен порядок, чем ближе к началу, тем больше приоритет для выбора при сравнение 
    d_min - первое значение из группы, берем как минимум по которому будет присваиться качество Диаметр<
    header_def - заголовки браков для формирования в excel
    sort_matching - сопоставление значений regsort индексам header_def для последующей обработке
    alarms - сигналы, оповестить, что-то не так
    """

    Diametr_KM = 50 
    d_min = group[0]
    regsort = ("Металл", "Диаметр<{}".format(d_min), "Cx","L", "C", "D", "Dc", "KM", "E", "AB")
    header_def = ["Диаметр<{}".format(d_min), "Металл", "Гниль, сложная кривизна", "Итого, брак"]
    sort_matching = {
                    "Диаметр<{}".format(d_min): 0,
                    "Металл": 1,
                    "Cx": 2,
                    "Dc": 2,
                    }

    timbers6m = [["AB {}-{}, см".format(group[i], group[i+1]-1), 0, 0] if i < len(group) - 1 else ["АВ от {} см. в торце до {} см. в комле".format(group[i], Diametr_KM), 0, 0] for i in range(len(group))]
    timbers57m = [record[:] for record in timbers6m]
    timbers6m.append(['Итого:', 0, 0])
    timbers57m.append(['Итого:', 0, 0])
    timbers_def = [[s, 0, 0] for s in header_def]
    timbers_data = timbers
    timbers_log = [] 
    timbers1C = []
    alarms = []

    i = 0
    p = 2

    clear_timbers(timbers, alarms, regsort)
    #if N % 2 > 0: return {'timbers6m': timbers6m, 'timbers57m': timbers57m, 'timbers_def': timbers_def, 'timbers1C': timbers1C, 'timbers_data': timbers_data, 'timbers_log': timbers_log}
    N = len(timbers)

    while i < N:
        sort = assign_sort(timbers[i][2], timbers[i+1][2], regsort)
        diametr = (min(timbers[i][4], timbers[i+1][4], d_min = d_min)) 
        timber1 = [timbers[i][0], timbers[i][2], timbers[i][5], timbers[i][4]]
        timber2 = [timbers[i+1][0], timbers[i+1][2], timbers[i+1][5], timbers[i+1][4]]
        #Кажется можно изящее 2 последующих if переопределяет сорт, если диаметр менее нужного или длина.
        #FIXME
        if diametr < d_min:
            sort = assign_sort(sort, "Диаметр<{}".format(d_min), regsort)
        
        length = check_length(timbers[i][5], timbers[i+1][5])
        if length == 5:
            #Переприсваиваем сорт L и делаем длину 6 для зачета объема
            sort = assign_sort(sort, "L", regsort)
            length = 6

        if sort == "AB" and length == 6 and diametr >= d_min:
            timbers6m[-1][2] += volume6[diametr]
            timbers6m[-1][1] += 1
            add_to_table(diametr = diametr, group = group, table = timbers6m, volumes = volume6, length = length)
            add_to_1c(diametr = diametr, table = timbers1C, sort = sort, length = length)
            add_to_log(target = "6 метров AB",ind = i, sort = 'AB', length = 6, diametr = diametr, timbers_log = timbers_log, timber1 = timber1, timber2 = timber2)

        elif sort == "AB" and length == 5.7 and diametr >= d_min:
            timbers6m[-1][2] += volume57[diametr]
            timbers6m[-1][1] += 1
            add_to_table(diametr = diametr, group = group, table = timbers6m, volumes = volume57)
            add_to_1c(diametr = diametr, table = timbers1C, sort = sort, length = 5.7)
            add_to_log(target = "5.7 метров AB",ind = i, sort = 'AB', length = 5.7, diametr = diametr, timbers_log = timbers_log, timber1 = timber1, timber2 = timber2)

        elif (sort == "AB") and (length == 6.5) and (diametr >= d_min):
            timbers6m[-1][2] += volume57[diametr]
            timbers6m[-1][1] += 1
            add_to_table(diametr = diametr, group = group, table = timbers6m, volumes = volume57)
            add_to_1c(diametr = diametr, table = timbers1C, sort = sort, length = 5.7)
            add_to_log(target = "6.2+ метров AB",ind = i, sort = 'AB', length = 5.7, diametr = diametr, timbers_log = timbers_log, timber1 = timber1, timber2 = timber2)

        elif sort == "E" and length == 6 and diametr >= d_min:
            timbers57m[-1][2] += volume6[diametr]
            timbers57m[-1][1] += 1
            add_to_table(diametr = diametr, group = group, table = timbers57m, volumes = volume6, length = length)
            add_to_1c(diametr = diametr, table = timbers1C, sort = sort, length = length)
            add_to_log(target = "6 метров E в 3ий сорт",ind = i, sort = 'E', length = 6, diametr = diametr, timbers_log = timbers_log, timber1 = timber1, timber2 = timber2)

        elif sort == "D" and length == 6 and diametr >= d_min:
            timbers57m[-1][2] += volume6[diametr]
            timbers57m[-1][1] += 1
            add_to_table(diametr = diametr, group = group, table = timbers57m, volumes = volume6, length = length)
            add_to_1c(diametr = diametr, table = timbers1C, sort = sort, length = length)
            add_to_log(target = "6 метров D в 3ий сорт",ind = i, sort = 'D', length = 6, diametr = diametr, timbers_log = timbers_log, timber1 = timber1, timber2 = timber2)
      
        elif sort == "D" and length == 5.7 and diametr >= d_min:
            timbers57m[-1][2] += volume57[diametr]
            timbers57m[-1][1] += 1
            add_to_table(diametr = diametr, group = group, table = timbers57m, volumes = volume57, length = length)
            add_to_1c(diametr = diametr, table = timbers1C, sort = sort, length = length)
            add_to_log(target = "5.7 метра D в 3ий сорт", ind = i, sort = 'D', length = length, diametr = diametr, timbers_log = timbers_log, timber1 = timber1, timber2 = timber2)

        elif sort == "C" and length == 6 and diametr >= d_min:
            timbers57m[-1][2] += volume6[diametr]
            timbers57m[-1][1] += 1
            add_to_table(diametr = diametr, group = group, table = timbers57m, volumes = volume6, length = length)
            add_to_1c(diametr = diametr, table = timbers1C, sort = sort, length = length)
            add_to_log(target = "6 метров C в 3ий сорт",ind = i, sort = 'C', length = 6, diametr = diametr, timbers_log = timbers_log, timber1 = timber1, timber2 = timber2)

        elif sort == "KM" and length == 6 and diametr >= d_min:
            timbers57m[-1][2] += volume6[diametr]
            timbers57m[-1][1] += 1
            add_to_table(diametr = diametr, group = group, table = timbers57m, volumes = volume6, length = length)
            add_to_1c(diametr = diametr, table = timbers1C, sort = sort, length = length)
            add_to_log(target = "6 метров KM в 3ий сорт",ind = i, sort = 'KM', length = 6, diametr = diametr, timbers_log = timbers_log, timber1 = timber1, timber2 = timber2)

        elif sort == "KM" and length == 5.7 and diametr >= d_min:
            timbers57m[-1][2] += volume57[diametr]
            timbers57m[-1][1] += 1
            add_to_table(diametr = diametr, group = group, table = timbers57m, volumes = volume57, length = length)
            add_to_1c(diametr = diametr, table = timbers1C, sort = sort, length = length)
            add_to_log(target = "5.7 метров KM в 3ий сорт",ind = i, sort = 'KM', length = 6, diametr = diametr, timbers_log = timbers_log, timber1 = timber1, timber2 = timber2)

        elif sort == "L" and length == 6 and diametr >= d_min:
            timbers57m[-1][2] += volume3[diametr]
            timbers57m[-1][1] += 1
            add_to_table(diametr = diametr, group = group, table = timbers57m, volumes = volume3, length = 3)
            add_to_1c(diametr = diametr, table = timbers1C, sort = sort, length = 3)
            add_to_log(target = "L в 3ий сорт объем, как 3 метра",ind = i, sort = 'L', length = 3, diametr = diametr, timbers_log = timbers_log, timber1 = timber1, timber2 = timber2)

        else:
            add_to_table_def(diametr = diametr, sort = sort, table = timbers_def, regsort = regsort, header_def = header_def,  sort_matching =  sort_matching, length = length, volume6 = volume6, volume57 = volume57)
            add_to_1c(diametr = diametr, table = timbers1C, sort = sort, length = length)
            add_to_log(target = "{1}м.  сорт:{0}".format(sort, length),ind = i, sort = sort, length = length, diametr = diametr, timbers_log = timbers_log, timber1 = timber1, timber2 = timber2)
        
        i += p

    else:
        timbers1C = sorted(counting_table(timbers1C), key=lambda rows: (rows[0], -rows[1], -rows[2]), reverse=True)

    return {'timbers6m': timbers6m, 'timbers57m': timbers57m, 'timbers_def': timbers_def, 'timbers1C': timbers1C, 'timbers_data': timbers_data, 'timbers_log': timbers_log, 'alarms': alarms}

def clear_timbers(timbers, alarms, regsort, d_min = 10, length_min = 2):
    #Функция очистки массива от лишнего,
    #Длины короче
    #Диаметра
    #Проверка на четность для 6 метров
    for t in timbers:
        if t[4] < d_min:
            timbers.remove(t)
            alarms.append("Очищено от бревен №{} по диаметру {}".format(t[0], t[4]))
        if float(t[5].replace(',', '.')) < length_min:
            timbers.remove(t)
            alarms.append("Очищено от бревен №{} по длине {}".format(t[0], t[5]))
        if t[2] not in regsort:
            alarms.append('Сорт "{}" не зарегистрированн в системе и заменен на сорт "Металл" бревно под номером {}'.format(t[2], t[0]))
    if len(timbers) % 2:
        timbers.pop()
        alarms.append("Удалено последнее бревно, т.к. кол-во не четное")


def add_to_log(target='Добавлено в список списков', ind = 1, sort = 'AB', length = 9, diametr = 69, timbers_log = None, **timbers):
    """ На входе спис
    timber1 - бревно 1
    timber2 - бревно 2
    target - куда добавляем
    Добавляем в logtable запись [target, # в 6ке, Сорт, Длина зачет, Диаметр, № 3м 1 часть, Сорт, Длина, Диаметр, № 3 часть 2, Сорт, Длина, Диаметр]
    Сделано через функцию для дальнешнего возможно преобразования. 
    """
    timbers_log.append([target, (ind//2+1), sort, length, diametr, *timbers['timber1'], *timbers['timber2']])

    

def add_to_1c(diametr=69, sort='AB', length=6, table = None):
    """Формируем список для таблицы 1C.
    Можно было обойтись без него, но в дальнейшем может пригодится.
    Удобно что тут контролируются все добавления в таблицу
    Через функцию затем, чтобы если что делать подмены и изменять данные если необходимо 
    """
    table.append([sort, length, diametr])

def counting_table(table):
    """Быстрее и проще пересчитать через collections counter, для тренировки пойдет
    На входе таблица для 1С в формате [сорт, длина, диаметр]
    B - пустой список для не повторябщихся бревен
    С - пустой список для кол-ва
    Далее через прогон списка и подссчет через метод count() заполняем B и С длина должна быть одинакова
    Делаем слияние B и С и возвращаем table2 
    """

    B = []
    C = []
    for t in table:
        if t not in B:
            B.append(t)
            C.append([table.count(t)])
    table2 = [(B[x]+C[x]) for x in range(len(B))]

    return table2

def min(a, b, d_min=14):
    """ Поиск минимального диаметра
    Добавить отбраковку по диаметру менее d_min, т.е. переопределить сорт
    """
    
    if a > b:
        return b 
    return a

def assign_sort(sort1, sort2, regsort):
    """Присвоить нужный сорт c проверкой на вхождение
    Присваивается тот сорт который ближе к началу списка.
    Поменяв местами изменится приоритет присваивания
    regsort - список сортов.

    """
    
    sort = "Металл"
    #FIXME
    #Проверка на вхождение в список и если OK начать проверку сортов.
    if sort1 in regsort and sort2 in regsort: 
        inx1 = regsort.index(sort1)
        inx2 = regsort.index(sort2)
        if inx1 < inx2:
            sort = regsort[inx1]
        else:
            sort = regsort[inx2]
        return sort
        
    return sort

def check_length(l1:str, l2:str):
    """Функция проверки длины, раскидываем в <5.7, 5.7, 6.05 
    notch - величина пропила, надо еще сорт переопределить, если короткое, но не выше

    """
    l1 = float(l1.replace(',','.'))
    l2 = float(l2.replace(',','.'))
    notch = 0.02
    sumlength = l1 + l2 + notch
    if sumlength <= 5.7:
        l = 5
    elif (sumlength > 5.7 and sumlength < 6.02):
        l = 5.7
    elif sumlength > 6.20:
        l = 6.5
    else:
        l = 6

    return l

def add_to_table(diametr=0, table=None, group=None, volumes=None, length=6):
    """Функция вычисляет объем на основании диаметра бревна и ведет подсчет количества.
    Добавляет в необходимый список в списке table кол-во и суммирует
    Аргументы функции: 
    diamert - диаметр бревна в см
    table - список ...
    group - группы диаметров
    volumes - словарь объемов ключ диаметр, значение объем
    Если диаметр меньше, то маркер кол-во в excel 9999
    Потом проверяем если больше максимального числа кладем в последнию ячейку
    Иначе проверяем интервал 
    
    """
    group = group or [10, 20, 40, 80]
    if diametr < group[0]:
       table[1][1] += 9999
    elif diametr >= group[-1]:
        table[len(group)-1][1] += 1
        table[len(group)-1][2] += volumes[diametr]
    else:
        for g in range(len(group)-1):
            if diametr >= (group[g]) and diametr < group[g+1]:
                table[g][1] += 1
                table[g][2] += volumes[diametr]


def add_to_table_def(diametr, sort, table, regsort, header_def, sort_matching, length = 6, volume6 = None, volume57 = None):
    """Функция заполняет таблицу с теми сортами которые попали в брак.
    Зачет объема будем делать по смотря какая длина 
    Для длин 5.7 и 6 разный объем

    """
    if length == 6:
        volumes = volume6
    else:
        volumes = volume57

    table[sort_matching[sort]][1] += 1 
    table[sort_matching[sort]][2] += volumes[diametr]
    table[-1][1] += 1 
    table[-1][2] += volumes[diametr]

def data_cleaning(basic_info):
    """Перерабатываем входный данные с парсера для формирования данных таблицы Excel
    """
    act_num = basic_info['act_num']
    act_num = re.search(r'\d+', act_num).group()
    basic_info['act_num'] = act_num

    date_first = basic_info['date_first']
    date_first = re.search(r'\d{2}.\d{2}.\d{4}', date_first).group()
    basic_info['date_first'] = date_first

    date_last = basic_info['date_last']
    date_last = re.search(r'\d{2}.\d{2}.\d{4}', date_last).group()
    basic_info['date_last'] = date_last

    return basic_info

def export_in_excel(template="app/doc_template/template-akt-2021-09.xlsx", **tables):
    """Заполнение шаблона excel
    """
    timbers1C = tables['timbers1C']
    timbers_data = tables['timbers_data']
    timbers_log = tables['timbers_log']
    timbers6m  = tables['timbers6m']
    timbers57m = tables['timbers57m']
    timbers_def = tables['timbers_def']
    alarms = tables['alarms']
    basic_info = tables['basic_info']
    wb = Workbook()
    template = template
    wb = load_workbook(template)
    
    ws1 = wb['акт1с_приемка']
    double = Side(border_style="thin", color="000000")
    for row in range(0, len(timbers1C)):
        for col in range(0, len(timbers1C[row])):
            ws1.cell(column = col+1, row = row+2, value=timbers1C[row][col])
    ws1.cell(column = 4, row = len(timbers1C)+2, value = '=SUM(D{}:D{})'.format(2, len(timbers1C)+1))

    
    ws1 = wb['данные']
    for row in range(0, len(timbers_data)):
        for col in range(0, len(timbers_data[row])):
            ws1.cell(column = col+1, row = row+17, value=timbers_data[row][col])
    #Блок основная информация по акту
    ws1.cell(column = 3, row = 3, value = basic_info['postavshik'])
    ws1.cell(column = 3, row = 4, value = basic_info['date_in'])
    ws1.cell(column = 3, row = 5, value = basic_info['place'])
    ws1.cell(column = 3, row = 6, value = basic_info['mark'])
    ws1.cell(column = 3, row = 7, value = basic_info['car_num'])
    ws1.cell(column = 3, row = 8, value = basic_info['trailer_num'])
    ws1.cell(column = 3, row = 9, value = basic_info['act_num'])
    ws1.cell(column = 3, row = 10, value =basic_info['operator'])
    ws1.cell(column = 3, row = 11, value =basic_info['date_first'])
    ws1.cell(column = 3, row = 12, value =basic_info['date_last'])

    ws1 = wb['log']
    p = 0
    for row in range(0, len(alarms)):
        ws1.cell(column=1, row=row+3, value=alarms[row])
        p += 1

    for row in range(0, len(timbers_log)):
        for col in range(0, len(timbers_log[row])):
            ws1.cell(column = col+1, row = row+3+p, value=timbers_log[row][col])

    ws1 = wb['акт приемки']
    #Основная информация на листе акт приемки
    ws1.cell(column=2, row=1, value = basic_info['act_num'])
    ws1.cell(column=2, row=2, value = basic_info['date_in'])
    ws1.cell(column=2, row=3, value = basic_info['date_first'])
    ws1.cell(column=2, row=4, value = basic_info['postavshik'])
    ws1.cell(column=2, row=5, value = basic_info['place'])
    ws1.cell(column=2, row=6, value = basic_info['mark'])
    ws1.cell(column=2, row=7, value = basic_info['car_num'])
    ws1.cell(column=2, row=8, value = basic_info['trailer_num'])

    for row in range(0, len(timbers6m)):
        for col in range(0, len(timbers6m[row])):
            ws1.cell(column = col+1, row = row+11, value=timbers6m[row][col])

    for row in range(0, len(timbers57m)):
        for col in range(0, len(timbers57m[row])):
            ws1.cell(column = col+1, row = row+19, value=timbers57m[row][col])

    for row in range(0, len(timbers_def)):
        for col in range(0, len(timbers_def[row])):
            ws1.cell(column = col+1, row = row+27, value=timbers_def[row][col])
    if len(alarms) != 0:
        ws1.cell(column = 1, row = 26, value="Во вкладке log есть примечания")

    filename = "app/tmp_files/akt/akt.xlsx"    
    
    try:
        wb.save((filename))
    except (OSError, IOError):
        print ("Файл занят")

    return filename
 


def test_proccessor():
    lst_timber = [ 
    ['1 ','Сосна ','Металл','249',25, '3,08', '0,170'], 
    ['2 ','Сосна ','AB','270',28, '3,08', '0,200'], 
    ['3 ','Сосна ','AB','261',26, '3,04', '0,185'], 
    ['4 ','Сосна ','Металл','241',24, '3,05', '0,157'], 
    ['5 ','Сосна ','AB','321',32, '3,04', '0,280'], 
    ['6 ','Сосна ','Cx','340',34, '3,02', '0,310'], 
    ['7 ','Сосна ','KM','285',29, '3,05', '0,230'], 
    ['8 ','Сосна ','AB','268',27, '3,04', '0,200'], 
    ['9 ','Сосна ','KM','277',28, '3,05', '0,220'], 
    ['10','Сосна', 'AB', '271' ,27, '3,03', '0,200'] ,  
    ['11','Сосна', 'AB', '281' ,28, '3,04', '0,220'] ,   
    ['12','Сосна', 'AB', '236' ,24, '2,65', '0,157'] ,  
    ['13','Сосна', 'AB', '246' ,13, '3,04', '0,170'] ,  
    ['14','Сосна', 'AB', '273' ,27, '3,06', '0,200'] ,  
    ['15','Сосна', 'C', '238' ,24, '3,03', '0,157'] ,   
    ['16','Сосна', 'AB', '271' ,27, '3,07', '0,200'] ,  
    ['17','Сосна', 'AB', '217' ,22, '3,04', '0,130'] ,  
    ['18','Сосна', 'AB', '196' ,20, '3,01', '0,107'] ,  
    ['19','Сосна', 'AB', '385' ,39, '3,03', '0,410'] ,  
    ['20','Сосна', 'AB', '394' ,39, '3,20', '0,410'] ,  
    ['21','Сосна', 'AB', '236' ,24, '2,20', '0,157'] ,  
    ['22','Сосна', 'AB', '259' ,26, '3,08', '0,185'] ,  
    ['23','Сосна', 'AB', '224' ,14, '3,06', '0,130'] ,  
    ['24','Сосна', 'AB', '192' ,19, '3,03', '0,096'] ,  
    ['25','Сосна', 'D', '385' ,29, '3,03', '0,410'] ,   
    ['26','Сосна', 'AB', '368' ,27, '3,10', '0,370'] ,  
    ['27','Сосна', 'Dc', '240' ,24, '3,04', '0,157'] ,
    ['28','Сосна', 'AB', '212' ,32, '3,07', '0,118'] ,  
    ['29','Сосна', 'AB', '212' ,14, '3,07', '0,118'] ,
    ['30','Сосна', 'D', '192' ,19, '3,05', '0,096'] ,
    ['31','Сосна', 'AB', '192' ,19, '3,05', '0,096'] ,
    ['32','Сосна', 'AB', '192' ,19, '2,90', '0,096'] ,
    ['33','Сосна', 'AB', '192' ,38, '3,05', '0,096'] , 
    ['34','Сосна', 'E', '192' ,29, '3,05', '0,096'] , 
    ['35' ,'Сосна','AB' ,'463' ,46 ,'3,03' ,'3,00' ,'0,570'],
    ['36' ,'Сосна','AB' ,'458' ,46 ,'1,00' ,'0,00' ,'0,000'],
    ['37' ,'Сосна','AB' ,'432' ,43 ,'3,00' ,'0,00' ,'0,000'] ,
    ['38' ,'Сосна','AB' ,'278' ,28 ,'3,03 ','3,00' ,'0,220'],
    ['39','Сосна', 'AB', '192' ,19, '3,05', '0,096'] ,
    ['40','Сосна', 'AB', '192' ,19, '2,90', '0,096'] 


    ]
    lst_timber1= [ 
    ['1','Сосна', 'AB', '192' ,36, '3,05', '0,096'] , 
    ['2','Сосна', 'AB', '192' ,36, '3,05', '0,096'] , 
    ['1','Сосна', 'AB', '192' ,35, '3,05', '0,096'] , 
    ['2','Сосна', 'AB', '192' ,35, '3,05', '0,096'] ,
    ['1','Сосна', 'AB', '192' ,34, '3,05', '0,096'] , 
    ['2','Сосна', 'AB', '192' ,34, '3,05', '0,096'] ,
    ['1','Сосна', 'AB', '192' ,33, '3,05', '0,096'] , 
    ['2','Сосна', 'AB', '192' ,33, '3,05', '0,096'],
    ['2','Сосна', 'AB', '192' ,32, '3,05', '0,096'],
    ['2','Сосна', 'AB', '192' ,32, '3,05', '0,096'],
    ['2','Сосна', 'AB', '192' ,31, '3,05', '0,096'],
    ['2','Сосна', 'AB', '192' ,31, '3,05', '0,096'],
    ]
    basic_info = {
                'postavshik': 'Арелан',
                'date_in': '22.01.2011',
                'place': 'Волхов',
                'mark': 'Интер',
                'car_num': 'С 645 ЕР 47' ,
                'trailer_num': 'АР 2567 47' ,
                'act_num': 'АКТ№715 ',
                'operator': 'Володащик Д.В. ' ,
                'date_first': '27.07.2021 13:47:37' ,
                'date_last': '27.07.2021 14:59:59 ',

    }
    x = main(lst_timber, group = [18, 20, 22, 24, 26, 35])
    export_in_excel(template="template-akt-2021-09.xlsx", 
                timbers1C = x['timbers1C'], 
                timbers_data = x['timbers_data'],
                timbers_log = x['timbers_log'],
                timbers_def = x['timbers_def'],
                timbers57m = x['timbers57m'],
                timbers6m = x['timbers6m'],
                alarms = x['alarms'],
                basic_info = data_cleaning(basic_info))



   #export_in_excel(template="template-akt-2021-07.xlsx", 
   #                timbers1C = x['timbers1C'], 
   #                timbers_data = x['timbers_data'],
   #                timbers_log = x['timbers_log'],
   #                timbers_def = x['timbers_def'],
   #                timbers57m = x['timbers57m'],
   #                timbers6m = x['timbers6m'],
   #                alarms = x['alarms'],
   #                basic_info = data_cleaning(basic_info)
   #                )


def make_act(lst_timber, basic_info):
    x = main(lst_timber, group = [18, 24, 28, 32, 36])
    export_in_excel(template="template-akt-2021-07.xlsx", 
                    timbers1C = x['timbers1C'], 
                    timbers_data = x['timbers_data'],
                    timbers_log = x['timbers_log'],
                    timbers_def = x['timbers_def'],
                    timbers57m = x['timbers57m'],
                    timbers6m = x['timbers6m'],
                    alarms = x['alarms'],
                    basic_info = data_cleaning(basic_info)
                    )

if __name__ == '__main__':
    test_proccessor()

